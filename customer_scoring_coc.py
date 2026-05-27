import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys

sys.stdout = open(r'c:/Users/patrickshen/WorkBuddy/20260517005736/coc_result.txt', 'w', encoding='utf-8')

# ========== CONFIG ==========
RAW_PATH = r'C:/Users/patrickshen/Desktop/工作/CRM/20260522_07_COC.xlsx'
OUT_PATH = r'C:/Users/patrickshen/Desktop/工作/CRM/COC_客户画像评分结果_v1.xlsx'

W_TEU    = 0.3
W_CM     = 0.5
W_CM_OFF = 0.2

# ========== LOAD DATA ==========
raw = pd.read_excel(RAW_PATH)
print(f'Raw data: {len(raw)} rows, {raw["SHIPPER_CD"].nunique()} customers')
print(f'Columns: {raw.columns.tolist()}')

raw = raw.rename(columns={
    'YM': 'REV_YM',
    'TEU': 'TEU_GP',
    'CM': 'CM_GP'
})

raw = raw.dropna(subset=['SHIPPER_CD', 'TEU_GP', 'CM_GP'])
raw = raw[raw['TEU_GP'] > 0]
raw['REV_YM']    = pd.to_datetime(raw['REV_YM'], format='%Y%m', errors='coerce')
raw = raw.dropna(subset=['REV_YM'])
raw['LANE']      = raw['LANE'].astype(str).str.strip()
raw['SHIPPER_CD'] = raw['SHIPPER_CD'].astype(str).str.strip()
raw['POL_CD']    = raw['POL_CD'].astype(str).str.strip()
raw['POD_CD']    = raw['POD_CD'].astype(str).str.strip()  # [IMPORT-EXCLUDE] 需要POD_CD判断进出口

# ========== [IMPORT-EXCLUDE] 进出口分类 ==========
raw['IS_IMPORT'] = (~raw['POL_CD'].str.startswith('CN')) & (raw['POD_CD'].str.startswith('CN'))
n_import = raw['IS_IMPORT'].sum()
print(f'\n=== Import/Export classification ===')
print(f'  Import records (POL!=CN & POD=CN): {n_import} ({n_import/len(raw)*100:.1f}%)')
print(f'  Non-import records: {len(raw) - n_import} ({(len(raw)-n_import)/len(raw)*100:.1f}%)')

# 用于评分的CM：进口记录的CM置0
raw['CM_GP_NOIMP'] = raw['CM_GP'].where(raw['IS_IMPORT'] == False, 0)

# ========== 1. IS_PEAK ==========
print('\n=== Seasonal Flag: CM/TEU annual average method ===')
print('  (Using CM excluding imports for lane baseline)')

lane_month_rate = (
    raw.groupby(['LANE','REV_YM'])
    .agg(MONTH_CM=('CM_GP_NOIMP','sum'), MONTH_TEU=('TEU_GP','sum'))  # [IMPORT-EXCLUDE]
    .reset_index()
)
lane_month_rate['MONTH_CM_PER_TEU'] = lane_month_rate['MONTH_CM'] / lane_month_rate['MONTH_TEU'].replace(0, np.nan)
lane_month_rate = lane_month_rate.drop(columns=['MONTH_CM','MONTH_TEU'])

lane_annual = (
    raw.groupby('LANE').agg(TOTAL_CM=('CM_GP_NOIMP','sum'), TOTAL_TEU=('TEU_GP','sum'))  # [IMPORT-EXCLUDE]
    .reset_index()
)
lane_annual['ANNUAL_AVG_CM_PER_TEU'] = lane_annual['TOTAL_CM'] / lane_annual['TOTAL_TEU']

lane_month_rate = lane_month_rate.merge(
    lane_annual[['LANE','ANNUAL_AVG_CM_PER_TEU']], on='LANE', how='left'
)
lane_month_rate['IS_PEAK'] = (
    lane_month_rate['MONTH_CM_PER_TEU'] >= lane_month_rate['ANNUAL_AVG_CM_PER_TEU']
).astype(int)

raw = raw.merge(lane_month_rate[['LANE','REV_YM','IS_PEAK']], on=['LANE','REV_YM'], how='left')

peak_cnt = (raw['IS_PEAK']==1).sum()
off_cnt   = (raw['IS_PEAK']==0).sum()
print(f'  Peak records: {peak_cnt}, Off-season records: {off_cnt}')

# ========== 2. CM_OFF 计算（新算法：分月评分 → 加权平均）==========
print('\n=== CM_OFF: monthly scoring + TEU-weighted average ===')
print('  (Excluding import records from off-season calculation)')

off_raw = raw[(raw['IS_PEAK']==0) & (raw['IS_IMPORT'] == False)].copy()  # [IMPORT-EXCLUDE]
print(f'  Off-season non-import records: {len(off_raw)}')

# 2.1 计算每月每条航线 CM/TEU（基准）
# off_raw已排除进口，CM_GP即为非进口CM
month_lane_median = (
    off_raw.groupby(['LANE','REV_YM'])
    .agg(MONTH_CM=('CM_GP','sum'), MONTH_TEU=('TEU_GP','sum'))
    .reset_index()
)
month_lane_median['MONTH_LANE_MEDIAN'] = month_lane_median['MONTH_CM'] / month_lane_median['MONTH_TEU'].replace(0, np.nan)
month_lane_median = month_lane_median[['LANE','REV_YM','MONTH_LANE_MEDIAN']]
print(f'  Month-lane median computed: {len(month_lane_median)} groups')

# 2.2 每个客户每条记录的月度评分
off_raw['MONTH_CM_PER_TEU'] = off_raw['CM_GP'] / off_raw['TEU_GP'].replace(0, np.nan)
off_raw = off_raw.merge(month_lane_median, on=['LANE','REV_YM'], how='left')
off_raw['MONTH_SCORE'] = off_raw['MONTH_CM_PER_TEU'] / off_raw['MONTH_LANE_MEDIAN'] * 100
off_raw['MONTH_SCORE'] = off_raw['MONTH_SCORE'].replace([np.inf,-np.inf], np.nan).fillna(0)

# 2.3 辅助函数
def teu_weighted_avg(df, group_cols):
    result = (
        df.groupby(group_cols)
        .apply(lambda g: np.average(g['MONTH_SCORE'], weights=g['TEU_GP']) if g['TEU_GP'].sum()>0 else 0)
        .reset_index()
    )
    result.columns = group_cols + ['CM_OFF_CM_PER_TEU']
    return result

# 2.4 各维度聚合
off_lane = teu_weighted_avg(off_raw, ['SHIPPER_CD','POL_CD','LANE'])
print(f'  L4 off records: {len(off_lane)}')

off_l3 = teu_weighted_avg(off_raw, ['SHIPPER_CD','POL_CD'])
print(f'  L3 off records: {len(off_l3)}')

off_l2 = teu_weighted_avg(off_raw, ['SHIPPER_CD','LANE'])
print(f'  L2 off records: {len(off_l2)}')

off_l1 = teu_weighted_avg(off_raw, ['SHIPPER_CD'])
print(f'  L1 off records: {len(off_l1)}')

# ========== 2.5 CM_OFF 硬门槛 ==========
OFF_THRESHOLD = 3
TEU_THRESHOLD = 50
print(f'\n=== CM_OFF hard threshold: OFF_RECORD>={OFF_THRESHOLD} AND OFF_TEU>={TEU_THRESHOLD} ===')

def apply_off_threshold_v2(df_off, group_cols):
    cnt = off_raw.groupby(group_cols).size().reset_index(name='OFF_CNT')
    teu = off_raw.groupby(group_cols)['TEU_GP'].sum().reset_index(name='OFF_TEU')
    df_off = df_off.merge(cnt, on=group_cols, how='left')
    df_off = df_off.merge(teu, on=group_cols, how='left')
    df_off['OFF_CNT'] = df_off['OFF_CNT'].fillna(0).astype(int)
    df_off['OFF_TEU'] = df_off['OFF_TEU'].fillna(0).astype(int)
    mask = (df_off['OFF_CNT'] < OFF_THRESHOLD) | (df_off['OFF_TEU'] < TEU_THRESHOLD)
    zeroed = mask.sum()
    df_off.loc[mask, 'CM_OFF_CM_PER_TEU'] = np.nan
    df_off.drop(columns=['OFF_CNT','OFF_TEU'], inplace=True)
    return df_off, zeroed

off_l1, zeroed_l1 = apply_off_threshold_v2(off_l1, ['SHIPPER_CD'])
off_l2, zeroed_l2 = apply_off_threshold_v2(off_l2, ['SHIPPER_CD','LANE'])
off_l3, zeroed_l3 = apply_off_threshold_v2(off_l3, ['SHIPPER_CD','POL_CD'])
off_lane, zeroed_l4 = apply_off_threshold_v2(off_lane, ['SHIPPER_CD','POL_CD','LANE'])

print(f'  L1 zeroed: {zeroed_l1} customers')
print(f'  L2 zeroed: {zeroed_l2} (SHIPPER_CD,LANE) groups')
print(f'  L3 zeroed: {zeroed_l3} (SHIPPER_CD,POL_CD) groups')
print(f'  L4 zeroed: {zeroed_l4} (SHIPPER_CD,POL_CD,LANE) groups')

# ========== FUNCTIONS ==========
def weighted_median(values):
    v = np.sort(values)[::-1].astype(float)
    if len(v) == 0: return 0
    cumsum = np.cumsum(v)
    total = cumsum[-1]
    if total <= 0: return 0
    idx = np.searchsorted(cumsum, total * 0.5)
    return v[min(idx, len(v)-1)]

def score_layer(agg, group_cols, col):
    if len(group_cols) == 0:
        wm = weighted_median(agg[col].values)
        agg[f'{col}_SCORE'] = agg[col] / wm * 100 if wm > 0 else 0.0
    else:
        wm_map = agg.groupby(group_cols)[col].apply(weighted_median).reset_index()
        wm_map.columns = group_cols + ['WM']
        agg = agg.merge(wm_map, on=group_cols, how='left')
        agg[f'{col}_SCORE'] = agg[col] / agg['WM'] * 100
        agg[f'{col}_SCORE'] = agg[f'{col}_SCORE'].replace([np.inf,-np.inf], np.nan).fillna(0)
        agg.drop(columns=['WM'], inplace=True)
    return agg

def score_by_simple_median(agg, group_cols, val_col, score_col):
    if len(group_cols) == 0:
        baseline = agg[val_col].median()
        agg[score_col] = agg[val_col] / baseline * 100 if baseline > 0 else 0.0
    else:
        med_map = agg.groupby(group_cols)[val_col].median().reset_index()
        med_map.columns = group_cols + ['BASELINE']
        agg = agg.merge(med_map, on=group_cols, how='left')
        agg[score_col] = agg[val_col] / agg['BASELINE'] * 100
        agg[score_col] = agg[score_col].replace([np.inf,-np.inf], np.nan).fillna(0)
        agg.drop(columns=['BASELINE'], inplace=True)
    return agg

# ========== 分位数分段映射 ==========
N_SEGMENTS = 20

def normalize_to_range(series, target_min=0, target_max=400):
    n_segs = N_SEGMENTS
    seg_width = (target_max - target_min) / n_segs
    result = pd.Series(np.nan, index=series.index, name=series.name, dtype=float)
    valid_mask = series.notna()
    valid = series[valid_mask]
    if len(valid) == 0:
        return pd.Series(100.0, index=series.index, name=series.name)
    if valid.max() == valid.min():
        return pd.Series(100.0, index=series.index, name=series.name)
    try:
        bins = pd.qcut(valid.rank(method='first'), n_segs, labels=False, duplicates='drop')
    except ValueError:
        s_min, s_max = valid.min(), valid.max()
        return (series - s_min) / (s_max - s_min) * (target_max - target_min) + target_min
    for seg_id in range(n_segs):
        seg_mask = bins == seg_id
        seg_values = valid[seg_mask]
        if len(seg_values) == 0:
            continue
        seg_lo = target_min + seg_id * seg_width
        seg_hi = target_min + (seg_id + 1) * seg_width
        v_min = seg_values.min()
        v_max = seg_values.max()
        if v_max == v_min:
            result[seg_values.index] = (seg_lo + seg_hi) / 2
        else:
            mapped = seg_lo + (seg_values - v_min) / (v_max - v_min) * seg_width
            result[seg_values.index] = mapped
    return result

def norm_nonzero(series):
    result = pd.Series(0.0, index=series.index, name=series.name)
    nonzero_mask = series != 0
    if nonzero_mask.sum() > 0:
        result.loc[nonzero_mask] = normalize_to_range(series.loc[nonzero_mask])
    return result

# ========== SHIPPER_NAME 规范化 ==========
print('\n=== Normalizing SHIPPER_NAME per SHIPPER_CD ===')
name_map = raw.groupby('SHIPPER_CD')['SHIPPER_NAME'].apply(
    lambda names: max(names.unique(), key=lambda x: len(str(x).replace('\xa0',' ').strip()))
).reset_index()
name_map.columns = ['SHIPPER_CD', 'SHIPPER_NAME_CANON']
multi_name = raw.groupby('SHIPPER_CD')['SHIPPER_NAME'].nunique()
multi_name_cnt = (multi_name > 1).sum()
print(f'  {raw["SHIPPER_CD"].nunique()} unique codes, {multi_name_cnt} codes with multiple names')

# ========== 3. 四个独立视图 ==========
print('\n=== Calculating 4 independent views (excluding imports from CM/CM_OFF) ===')

# ===== L2: 航线 =====
print('\n--- L2: By Lane ---')
l2 = raw.groupby(['SHIPPER_CD','LANE']).agg(
    TEU=('TEU_GP','sum'),
    CM=('CM_GP','sum'),            # [IMPORT-EXCLUDE] 原始CM（含进口），用于展示
    CM_NI=('CM_GP_NOIMP','sum')    # [IMPORT-EXCLUDE] 排除进口的CM，用于评分
).reset_index()
l2 = l2.merge(name_map, on='SHIPPER_CD', how='left')
l2.rename(columns={'SHIPPER_NAME_CANON':'SHIPPER_NAME'}, inplace=True)
l2['CM_PER_TEU'] = l2['CM_NI'] / l2['TEU'].replace(0, np.nan)  # [IMPORT-EXCLUDE]
l2 = l2.merge(off_l2[['SHIPPER_CD','LANE','CM_OFF_CM_PER_TEU']], on=['SHIPPER_CD','LANE'], how='left')
l2['CM_OFF_CM_PER_TEU'] = l2['CM_OFF_CM_PER_TEU'].replace([np.inf,-np.inf], np.nan)

# TEU评分（全量，含进口）
l2 = score_layer(l2, ['LANE'], 'TEU')

# CM评分：[IMPORT-EXCLUDE] 使用CM_NI
l2.loc[l2['CM_NI'] < 0, 'CM_PER_TEU'] = np.nan  # [IMPORT-EXCLUDE]
l2 = score_by_simple_median(l2, ['LANE'], 'CM_PER_TEU', 'CM_SCORE')
l2['CM_SCORE'] = l2['CM_SCORE'].replace([np.inf,-np.inf], 0).fillna(0)

# OffSeason评分：[IMPORT-EXCLUDE] 使用CM_NI判断
l2 = score_by_simple_median(l2, ['LANE'], 'CM_OFF_CM_PER_TEU', 'CM_OFF_SCORE')
l2['CM_OFF_SCORE'] = l2['CM_OFF_SCORE'].replace([np.inf,-np.inf], 0).fillna(0)
l2.loc[l2['CM_NI'] < 0, 'CM_OFF_SCORE'] = 0  # [IMPORT-EXCLUDE]

# L2标准化 + 总分
teu_norm    = normalize_to_range(l2['TEU_SCORE'])
cm_norm     = norm_nonzero(l2['CM_SCORE'])
cm_off_norm = norm_nonzero(l2['CM_OFF_SCORE'])
l2['TEU_NORM']    = teu_norm
l2['CM_NORM']     = cm_norm
l2['CM_OFF_NORM'] = cm_off_norm
l2['TOTAL'] = W_TEU * teu_norm + W_CM * cm_norm + W_CM_OFF * cm_off_norm
l2['RANK'] = l2.groupby('LANE')['TOTAL'].rank(ascending=False, method='min').astype(int)
l2 = l2.sort_values(['LANE','TOTAL'], ascending=[True, False]).reset_index(drop=True)
print(f'  {len(l2)} records across {l2["LANE"].nunique()} lanes')

# ===== L4: 口岸×航线 =====
print('\n--- L4: By Port×Lane ---')
l4 = raw.groupby(['SHIPPER_CD','POL_CD','LANE']).agg(
    TEU=('TEU_GP','sum'),
    CM=('CM_GP','sum'),            # [IMPORT-EXCLUDE] 原始CM
    CM_NI=('CM_GP_NOIMP','sum')    # [IMPORT-EXCLUDE] 排除进口的CM
).reset_index()
l4 = l4.merge(name_map, on='SHIPPER_CD', how='left')
l4.rename(columns={'SHIPPER_NAME_CANON':'SHIPPER_NAME'}, inplace=True)
l4['CM_PER_TEU'] = l4['CM_NI'] / l4['TEU'].replace(0, np.nan)  # [IMPORT-EXCLUDE]
l4 = l4.merge(off_lane[['SHIPPER_CD','POL_CD','LANE','CM_OFF_CM_PER_TEU']],
                on=['SHIPPER_CD','POL_CD','LANE'], how='left')
l4['CM_OFF_CM_PER_TEU'] = l4['CM_OFF_CM_PER_TEU'].replace([np.inf,-np.inf], np.nan)

# TEU评分
l4 = score_layer(l4, ['POL_CD','LANE'], 'TEU')

# CM评分：[IMPORT-EXCLUDE]
l4.loc[l4['CM_NI'] < 0, 'CM_PER_TEU'] = np.nan
l4 = score_by_simple_median(l4, ['POL_CD','LANE'], 'CM_PER_TEU', 'CM_SCORE')
l4['CM_SCORE'] = l4['CM_SCORE'].replace([np.inf,-np.inf], 0).fillna(0)

# OffSeason评分：[IMPORT-EXCLUDE]
l4 = score_by_simple_median(l4, ['POL_CD','LANE'], 'CM_OFF_CM_PER_TEU', 'CM_OFF_SCORE')
l4['CM_OFF_SCORE'] = l4['CM_OFF_SCORE'].replace([np.inf,-np.inf], 0).fillna(0)
l4.loc[l4['CM_NI'] < 0, 'CM_OFF_SCORE'] = 0

# L4标准化 + 总分
teu_norm    = normalize_to_range(l4['TEU_SCORE'])
cm_norm     = norm_nonzero(l4['CM_SCORE'])
cm_off_norm = norm_nonzero(l4['CM_OFF_SCORE'])
l4['TEU_NORM']    = teu_norm
l4['CM_NORM']     = cm_norm
l4['CM_OFF_NORM'] = cm_off_norm
l4['TOTAL'] = W_TEU * teu_norm + W_CM * cm_norm + W_CM_OFF * cm_off_norm
l4['RANK'] = l4.groupby(['POL_CD','LANE'])['TOTAL'].rank(ascending=False, method='min').astype(int)
l4 = l4.sort_values(['POL_CD','LANE','TOTAL'], ascending=[True, True, False]).reset_index(drop=True)
print(f'  {len(l4)} records across {l4.groupby(["POL_CD","LANE"]).ngroups} port×lane combos')

# ===== L1: 总体 =====
print('\n--- L1: Overall (CM_Score from L2 weighted average) ---')
l1 = raw.groupby(['SHIPPER_CD']).agg(
    TEU=('TEU_GP','sum'),
    CM=('CM_GP','sum'),            # [IMPORT-EXCLUDE] 原始CM
    CM_NI=('CM_GP_NOIMP','sum')    # [IMPORT-EXCLUDE] 排除进口的CM
).reset_index()
l1 = l1.merge(name_map, on='SHIPPER_CD', how='left')
l1.rename(columns={'SHIPPER_NAME_CANON':'SHIPPER_NAME'}, inplace=True)
l1 = l1.merge(off_l1[['SHIPPER_CD','CM_OFF_CM_PER_TEU']], on='SHIPPER_CD', how='left')
l1['CM_OFF_CM_PER_TEU'] = l1['CM_OFF_CM_PER_TEU'].replace([np.inf,-np.inf], np.nan)

# TEU评分（全量，含进口）
l1 = score_layer(l1, [], 'TEU')

# CM评分：从L2按CM占比加权汇总 [IMPORT-EXCLUDE] 使用CM_NI
l2_for_l1 = l2[['SHIPPER_CD','LANE','CM_NI','CM_SCORE']].copy()  # [IMPORT-EXCLUDE]
l2_pos = l2_for_l1[l2_for_l1['CM_NI'] > 0].copy()  # [IMPORT-EXCLUDE]
if len(l2_pos) > 0:
    cm_total = l2_pos.groupby('SHIPPER_CD')['CM_NI'].sum().reset_index(name='CM_POS_TOTAL')  # [IMPORT-EXCLUDE]
    l2_pos = l2_pos.merge(cm_total, on='SHIPPER_CD', how='left')
    l2_pos['CM_WEIGHT'] = l2_pos['CM_NI'] / l2_pos['CM_POS_TOTAL']  # [IMPORT-EXCLUDE]
    l2_pos['CM_SCORE_WEIGHTED'] = l2_pos['CM_SCORE'] * l2_pos['CM_WEIGHT']
    l1_cm = l2_pos.groupby('SHIPPER_CD')['CM_SCORE_WEIGHTED'].sum().reset_index(name='CM_SCORE')
    l1 = l1.merge(l1_cm, on='SHIPPER_CD', how='left')
else:
    l1['CM_SCORE'] = 0
l1['CM_SCORE'] = l1['CM_SCORE'].replace([np.inf,-np.inf], 0).fillna(0)
l1.loc[l1['CM_NI'] < 0, 'CM_SCORE'] = 0  # [IMPORT-EXCLUDE]

# OffSeason评分（独立计算）
l1 = score_by_simple_median(l1, [], 'CM_OFF_CM_PER_TEU', 'CM_OFF_SCORE')
l1['CM_OFF_SCORE'] = l1['CM_OFF_SCORE'].replace([np.inf,-np.inf], 0).fillna(0)
l1.loc[l1['CM_NI'] < 0, 'CM_OFF_SCORE'] = 0  # [IMPORT-EXCLUDE]

# L1标准化 + 总分
teu_norm    = normalize_to_range(l1['TEU_SCORE'])
cm_norm     = norm_nonzero(l1['CM_SCORE'])
cm_off_norm = norm_nonzero(l1['CM_OFF_SCORE'])
l1['TEU_NORM']    = teu_norm
l1['CM_NORM']     = cm_norm
l1['CM_OFF_NORM'] = cm_off_norm
l1['TOTAL'] = W_TEU * teu_norm + W_CM * cm_norm + W_CM_OFF * cm_off_norm
l1['RANK']  = l1['TOTAL'].rank(ascending=False, method='min').astype(int)
l1['TEU_RANK']    = l1['TEU_NORM'].rank(ascending=False, method='min').astype(int)
l1['CM_RANK']     = l1['CM_NORM'].rank(ascending=False, method='min').astype(int)
l1['CM_OFF_RANK'] = l1['CM_OFF_NORM'].rank(ascending=False, method='min').astype(int)
l1 = l1.sort_values('TOTAL', ascending=False).reset_index(drop=True)
print(f'  {len(l1)} customers')
print(l1.head(10)[['RANK','SHIPPER_CD','CM','CM_NI','TEU_SCORE','CM_SCORE','CM_OFF_SCORE','TOTAL']].to_string())

# ===== L3: 口岸 =====
print('\n--- L3: By Port (CM_Score from L4 weighted average) ---')
l3 = raw.groupby(['SHIPPER_CD','POL_CD']).agg(
    TEU=('TEU_GP','sum'),
    CM=('CM_GP','sum'),            # [IMPORT-EXCLUDE] 原始CM
    CM_NI=('CM_GP_NOIMP','sum')    # [IMPORT-EXCLUDE] 排除进口的CM
).reset_index()
l3 = l3.merge(name_map, on='SHIPPER_CD', how='left')
l3.rename(columns={'SHIPPER_NAME_CANON':'SHIPPER_NAME'}, inplace=True)
l3 = l3.merge(off_l3[['SHIPPER_CD','POL_CD','CM_OFF_CM_PER_TEU']], on=['SHIPPER_CD','POL_CD'], how='left')
l3['CM_OFF_CM_PER_TEU'] = l3['CM_OFF_CM_PER_TEU'].replace([np.inf,-np.inf], np.nan)

# TEU评分
l3 = score_layer(l3, ['POL_CD'], 'TEU')

# CM评分：从L4按CM占比加权汇总 [IMPORT-EXCLUDE]
l4_for_l3 = l4[['SHIPPER_CD','POL_CD','LANE','CM_NI','CM_SCORE']].copy()  # [IMPORT-EXCLUDE]
l4_pos = l4_for_l3[l4_for_l3['CM_NI'] > 0].copy()  # [IMPORT-EXCLUDE]
if len(l4_pos) > 0:
    cm_total_l3 = l4_pos.groupby(['SHIPPER_CD','POL_CD'])['CM_NI'].sum().reset_index(name='CM_POS_TOTAL')  # [IMPORT-EXCLUDE]
    l4_pos = l4_pos.merge(cm_total_l3, on=['SHIPPER_CD','POL_CD'], how='left')
    l4_pos['CM_WEIGHT'] = l4_pos['CM_NI'] / l4_pos['CM_POS_TOTAL']  # [IMPORT-EXCLUDE]
    l4_pos['CM_SCORE_WEIGHTED'] = l4_pos['CM_SCORE'] * l4_pos['CM_WEIGHT']
    l3_cm = l4_pos.groupby(['SHIPPER_CD','POL_CD'])['CM_SCORE_WEIGHTED'].sum().reset_index(name='CM_SCORE')
    l3 = l3.merge(l3_cm, on=['SHIPPER_CD','POL_CD'], how='left')
else:
    l3['CM_SCORE'] = 0
l3['CM_SCORE'] = l3['CM_SCORE'].replace([np.inf,-np.inf], 0).fillna(0)
l3.loc[l3['CM_NI'] < 0, 'CM_SCORE'] = 0  # [IMPORT-EXCLUDE]

# OffSeason评分
l3 = score_by_simple_median(l3, ['POL_CD'], 'CM_OFF_CM_PER_TEU', 'CM_OFF_SCORE')
l3['CM_OFF_SCORE'] = l3['CM_OFF_SCORE'].replace([np.inf,-np.inf], 0).fillna(0)
l3.loc[l3['CM_NI'] < 0, 'CM_OFF_SCORE'] = 0  # [IMPORT-EXCLUDE]

# L3标准化 + 总分
teu_norm    = normalize_to_range(l3['TEU_SCORE'])
cm_norm     = norm_nonzero(l3['CM_SCORE'])
cm_off_norm = norm_nonzero(l3['CM_OFF_SCORE'])
l3['TEU_NORM']    = teu_norm
l3['CM_NORM']     = cm_norm
l3['CM_OFF_NORM'] = cm_off_norm
l3['TOTAL'] = W_TEU * teu_norm + W_CM * cm_norm + W_CM_OFF * cm_off_norm
l3['RANK'] = l3.groupby('POL_CD')['TOTAL'].rank(ascending=False, method='min').astype(int)
l3 = l3.sort_values(['POL_CD','TOTAL'], ascending=[True, False]).reset_index(drop=True)
print(f'  {len(l3)} records across {l3["POL_CD"].nunique()} ports')

# ========== 4. 输出 Excel ==========
print('\n=== Writing Excel ===')
with pd.ExcelWriter(OUT_PATH, engine='openpyxl') as writer:
    l1_out = l1[['RANK','SHIPPER_CD','SHIPPER_NAME','TEU','CM','CM_NI',
                  'TEU_NORM','TEU_RANK',
                  'CM_NORM','CM_RANK',
                  'CM_OFF_NORM','CM_OFF_RANK',
                  'TOTAL']].copy()
    l1_out.columns = ['Rank','Code','Name','TEU','CM','CM_NI',
                        'TEU_Score','TEU_Rank',
                        'CM_Score','CM_Rank',
                        'OffSeason_Score','OffSeason_Rank',
                        'Total_Score']
    l1_out.to_excel(writer, sheet_name='L1_Overall', index=False)

    l2_out = l2[['LANE','RANK','SHIPPER_CD','SHIPPER_NAME','TEU','CM','CM_NI',
                  'TEU_NORM',
                  'CM_NORM',
                  'CM_OFF_NORM','TOTAL']].copy()
    l2_out.columns = ['Lane','Rank','Code','Name','TEU','CM','CM_NI',
                        'TEU_Score',
                        'CM_Score',
                        'OffSeason_Score','Total_Score']
    l2_out.to_excel(writer, sheet_name='L2_ByLane', index=False)

    l3_out = l3[['POL_CD','RANK','SHIPPER_CD','SHIPPER_NAME','TEU','CM','CM_NI',
                  'TEU_NORM',
                  'CM_NORM',
                  'CM_OFF_NORM','TOTAL']].copy()
    l3_out.columns = ['Port','Rank','Code','Name','TEU','CM','CM_NI',
                        'TEU_Score',
                        'CM_Score',
                        'OffSeason_Score','Total_Score']
    l3_out.to_excel(writer, sheet_name='L3_ByPort', index=False)

    l4_out = l4[['POL_CD','LANE','RANK','SHIPPER_CD','SHIPPER_NAME','TEU','CM','CM_NI',
                  'TEU_NORM',
                  'CM_NORM',
                  'CM_OFF_NORM','TOTAL']].copy()
    l4_out.columns = ['Port','Lane','Rank','Code','Name','TEU','CM','CM_NI',
                        'TEU_Score',
                        'CM_Score',
                        'OffSeason_Score','Total_Score']
    l4_out.to_excel(writer, sheet_name='L4_ByPortLane', index=False)

# Format Excel
wb = load_workbook(OUT_PATH)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')
thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(horizontal='right')
        ws.column_dimensions[col_letter].width = min(max_len+4, 30)
    ws.freeze_panes = 'A2'
wb.save(OUT_PATH)
print(f'Excel saved: {OUT_PATH}')

print('\nDONE')
sys.stdout.close()
